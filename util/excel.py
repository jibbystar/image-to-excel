# Import required libraries
import openpyxl
import openpyxl.styles as exstyles
import openpyxl.writer.excel as exwriter
import openpyxl.utils as utils

class ExcelDoc:
    "Writes and controls excel files."

    def __init__(self, resize_cells: bool = True) -> None:
        # Create the workbook
        self.workbook = openpyxl.Workbook()
        self.resize_cells = resize_cells

    def draw(self, rgba_map: list[list[tuple[int, int, int, int]]], cell_size_division_amount: int | float) -> None:
        "Draws the RGBA map provided onto the Excel active worksheet."

        sheet = self.workbook.active
        # sheet.add_image(...) IM JUST KIDDING
        y = 0
        x = 0
        for y_ in rgba_map:
            y += 1
            x = 0
            if self.resize_cells:
                # Resize each column
                sheet.column_dimensions[utils.get_column_letter(y)].width = 13.57 / cell_size_division_amount
            else:
                sheet.column_dimensions[utils.get_column_letter(y)].width /= cell_size_division_amount
            for x_ in y_:
                x += 1
                if self.resize_cells:
                    sheet.row_dimensions[x].height = 75.00 / cell_size_division_amount
                else:
                    sheet.row_dimensions[x].height /= cell_size_division_amount
                sheet.cell(y, x).fill = exstyles.fills.PatternFill("solid", exstyles.Color("{:02x}{:02x}{:02x}".format(*x_)))
                #print("{:02x}{:02x}{:02x}".format(*x_))

    def write_to(self, path: str):
        "Writes the Excel workbook to the path provided."
        
        try:
            exwriter.save_workbook(self.workbook, path)
        except KeyError as e:
            print(e.with_traceback())
